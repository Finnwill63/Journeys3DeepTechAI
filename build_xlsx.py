# -*- coding: utf-8 -*-
import csv, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

M=list(csv.reader(open('/root/work/_mentors.csv',encoding='utf-8')))
matches=json.load(open('/root/work/_matches.json',encoding='utf-8'))

EXP={'Fundraising & Investor Readiness':'C','Go-to-market, Sales & Business Development':'D',
 'Strategy & Business Model':'E','Team, Leadership & Operations':'F',
 'Corporate Partnerships & Ecosystems':'G','Marketing, Brand & Communications':'H',
 'Product & Design':'I','Finance, Legal & IP':'J'}

def mstatus(name,join):
    if name=='Will Cardwell': return 'Host'
    if join.startswith('Unfortunately'): return 'Declined'
    if join.startswith('Not'): return 'Maybe'
    return 'In'

mrows=[]
for r in M[1:]:
    if len(r)>16: r=r[:5]+[r[5]+' '+r[6]]+r[7:]  # fix embedded-quote row
    name=r[1].strip()
    join=r[2].strip()
    sect=r[3]
    d={'name':name,'status':mstatus(name,join),
       'exp':{lab for lab in EXP if lab in r[4]},
       'ai':'AI/ML & Data' in sect,'health':'Health & Bio' in sect,
       'climhw':('Energy, Climate & Sustainability' in sect or 'Hardware, Materials & Manufacturing' in sect),
       'soft':'Software / Cloud / Telecom' in sect,'cross':'work across sectors' in sect,
       'sectors':sect.replace('I work across sectors — don’t tie me to one','Cross-sector').strip(),
       'cap':(r[7].strip().split(',')[0] if r[7].strip() else ''),
       'cadence':r[8].strip(),'mode':r[9].strip(),
       'base':('Finland' if r[10].strip()=='Finland' else (r[11].strip() or ('Abroad' if r[10].strip() else ''))),
       'country':r[11].strip(),'demo':r[12].strip(),'focus':r[5].strip(),'notes':''}
    mrows.append(d)

# ---------- styles ----------
ARIAL='Arial'
INK='1A1A2E'; BLUE='2563A8'; LT='EAF1FA'; YEL='FFF6D6'; GREY='F4F5F9'
def F(**k): return Font(name=ARIAL, **k)
hdr_fill=PatternFill('solid',fgColor=BLUE)
hdr_font=F(bold=True,color='FFFFFF',size=10)
title_font=F(bold=True,size=16,color=INK)
sub_font=F(italic=True,size=10,color='6B6F80')
lab_font=F(bold=True,size=10,color=BLUE)
yfill=PatternFill('solid',fgColor=YEL)
ltfill=PatternFill('solid',fgColor=LT)
gfill=PatternFill('solid',fgColor=GREY)
thin=Side(style='thin',color='D8DBE6')
border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True,vertical='top')
ctr=Alignment(horizontal='center',vertical='center')

wb=Workbook()

# ================= START HERE =================
s=wb.active; s.title='Start Here'
s.sheet_view.showGridLines=False
s['A1']='Journeys 3 — Deep Tech & AI  ·  Mentor Matching Workbook'; s['A1'].font=F(bold=True,size=15,color=INK)
s['A2']='Internal — Will & Fabian'; s['A2'].font=F(italic=True,size=10,color='B0592B')
notes=[
 ('','' ),
 ('How this workbook is organised',''),
 ('One tab per venture (12)','Each shows the team’s stated needs and the top 3 mentor matches, ranked by fit (expertise ↔ needs + sector). Details (status, capacity, sector, base) auto-fill from the Mentors tab.'),
 ('Mentors tab','The full searchable directory of all mentor sign-ups. Use the filter arrows on the header row to search by expertise (Y = yes), sector, capacity, base, etc.'),
 ('Adding ad-hoc mentors','Add anyone new at the bottom of the Mentors tab (yellow rows). Once added, type their name into a venture tab’s “Add other / ad-hoc” rows and their details auto-fill.'),
 ('Assigning','On each venture tab, use the yellow “Assign?” column to mark your pick (Assigned / Backup / No) and jot notes.'),
 ('',''),
 ('Scope','14 founder responses across 11 teams; White Hat from assessment (no survey). 21 mentors. Pool for matches excludes Will (host) and Samuli (declined). “Maybe” = not yet confirmed (Hanna, Anja).'),
]
r=3
for a,b in notes:
    s.cell(r,1,a).font=lab_font if b else F(bold=True,size=11,color=INK)
    if b: s.cell(r,2,b).font=F(size=10,color=INK); s.cell(r,2).alignment=wrap
    r+=1
s.column_dimensions['A'].width=26; s.column_dimensions['B'].width=95
for rr in range(3,r): s.row_dimensions[rr].height=30

# ================= MENTORS =================
mt=wb.create_sheet('Mentors')
mt.sheet_view.showGridLines=False
mt['A1']='Mentor Directory — searchable'; mt['A1'].font=title_font
mt['A2']='Filter with the header arrows (row 4). Expertise columns: Y = yes. Add ad-hoc mentors in the yellow rows at the bottom — they become searchable here and usable on the venture tabs.'
mt['A2'].font=sub_font; mt['A2'].alignment=Alignment(wrap_text=False)
HEAD=['Name','Status','Fundraising','GTM / Sales','Strategy / BM','Team / Ops','Partnerships',
 'Marketing','Product','Finance/Legal/IP','AI / Data','Health','Climate / HW','Software','Cross-sector',
 'Sectors (detail)','Capacity','Cadence','Online/In-person','Based','Country','Demo Day','Wants to focus on','Notes']
HR=4
for c,h in enumerate(HEAD,1):
    cell=mt.cell(HR,c,h); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=Alignment(wrap_text=True,vertical='center',horizontal='center'); cell.border=border
rr=HR+1
def yn(b): return 'Y' if b else ''
for d in mrows:
    vals=[d['name'],d['status'],yn('Fundraising & Investor Readiness' in d['exp']),
      yn('Go-to-market, Sales & Business Development' in d['exp']),yn('Strategy & Business Model' in d['exp']),
      yn('Team, Leadership & Operations' in d['exp']),yn('Corporate Partnerships & Ecosystems' in d['exp']),
      yn('Marketing, Brand & Communications' in d['exp']),yn('Product & Design' in d['exp']),
      yn('Finance, Legal & IP' in d['exp']),yn(d['ai']),yn(d['health']),yn(d['climhw']),yn(d['soft']),yn(d['cross']),
      d['sectors'],d['cap'],d['cadence'],d['mode'],d['base'],d['country'],d['demo'],d['focus'],d['notes']]
    for c,v in enumerate(vals,1):
        cell=mt.cell(rr,c,v); cell.font=F(size=9); cell.border=border; cell.alignment=Alignment(vertical='center',wrap_text=(c in(16,23)))
        if c in (3,4,5,6,7,8,9,10,11,12,13,14,15) and v=='Y': cell.alignment=ctr; cell.font=F(size=9,bold=True,color=BLUE)
        elif c in (3,4,5,6,7,8,9,10,11,12,13,14,15): cell.alignment=ctr
        if d['status']=='Declined': cell.font=F(size=9,color='9A9DB0',italic=True)
    rr+=1
DATA_END=rr-1
# ad-hoc example + blank rows
ex=['Jane Example','Ad-hoc','Y','','Y','','','','','','Y','','','','', 'AI/ML & Data','One','','Online only','Abroad','—','—','Domain expert intro','example row — overwrite']
for c,v in enumerate(ex,1):
    cell=mt.cell(rr,c,v); cell.font=F(size=9,italic=True,color='9A9DB0'); cell.border=border; cell.alignment=Alignment(vertical='center',wrap_text=(c in(16,23)))
    if c in range(3,16): cell.alignment=ctr
rr+=1
ADHOC_START=rr
for _ in range(15):
    for c in range(1,25):
        cell=mt.cell(rr,c,None); cell.border=border; cell.font=F(size=9)
        if c==1: cell.fill=yfill
    rr+=1
LASTROW=rr-1
mt.auto_filter.ref=f'A{HR}:X{LASTROW}'
mt.freeze_panes='A5'
widths=[22,9,8,8,8,7,8,8,7,9,6,6,7,7,7,34,9,16,14,12,14,20,40,26]
for i,w in enumerate(widths,1): mt.column_dimensions[get_column_letter(i)].width=w
# status dropdown
dv=DataValidation(type='list',formula1='"In,Maybe,Declined,Host,Ad-hoc"',allow_blank=True)
mt.add_data_validation(dv); dv.add(f'B{HR+1}:B{LASTROW}')

# ================= COMPANY TABS =================
assign_dv_formula='"Assigned,Backup,No"'
for t in matches:
    nm=t['team'][:31]
    cs=wb.create_sheet(nm)
    cs.sheet_view.showGridLines=False
    cs['A1']=t['team']; cs['A1'].font=title_font
    cs['A2']=t['blurb']; cs['A2'].font=sub_font
    cs['A3']='Team’s stated needs:'; cs['A3'].font=lab_font
    cs['B3']=' · '.join(t['needtags']); cs['B3'].font=F(size=10,color=INK)
    CH=['Rank','Mentor','Why this is a fit','Status','Capacity','Sector','Based','Assign?','Notes']
    hrow=5
    for c,h in enumerate(CH,1):
        cell=cs.cell(hrow,c,h); cell.fill=hdr_fill; cell.font=hdr_font; cell.border=border; cell.alignment=Alignment(wrap_text=True,vertical='center',horizontal='center')
    row=hrow+1
    def lookup(colletter,namecell):
        return f'=IFERROR(INDEX(Mentors!${colletter}:${colletter},MATCH(${namecell},Mentors!$A:$A,0)),"")'
    def write_match(rank,name,why):
        cs.cell(row,1,rank).alignment=ctr; cs.cell(row,1).font=F(bold=True,color=BLUE)
        cs.cell(row,2,name).font=F(bold=True,size=10)
        cs.cell(row,3,why).font=F(size=9); cs.cell(row,3).alignment=wrap
        nc=f'B{row}'
        cs.cell(row,4,lookup('B',nc)).font=F(size=9)
        cs.cell(row,5,lookup('Q',nc)).font=F(size=9); cs.cell(row,5).alignment=ctr
        cs.cell(row,6,lookup('P',nc)).font=F(size=9); cs.cell(row,6).alignment=wrap
        cs.cell(row,7,lookup('T',nc)).font=F(size=9)
        cs.cell(row,8,None).fill=yfill
        cs.cell(row,9,None).fill=yfill
        for c in range(1,10): cs.cell(row,c).border=border
    for i,m in enumerate(t['matches'],1):
        write_match(i,m['name'],m['why']); row+=1
    # separator
    row+=1
    cs.cell(row,1,'Add other / ad-hoc candidates').font=lab_font
    cs.cell(row,2,'Type a mentor name (must exist on the Mentors tab) — their details auto-fill. Add brand-new people on the Mentors tab first.').font=F(size=9,italic=True,color='6B6F80')
    cs.cell(row,2).alignment=Alignment(wrap_text=False)
    row+=1
    # example row
    cs.cell(row,1,'—').alignment=ctr
    cs.cell(row,2,'Jane Example').font=F(size=10,italic=True,color='9A9DB0')
    cs.cell(row,3,'why you’re considering them').font=F(size=9,italic=True,color='9A9DB0'); cs.cell(row,3).alignment=wrap
    for cl,col in [('B',4),('Q',5),('P',6),('T',7)]:
        cs.cell(row,col,lookup(cl,f'B{row}')).font=F(size=9,italic=True,color='9A9DB0')
    for c in range(1,10): cs.cell(row,c).border=border
    row+=1
    adhoc_first=row
    for _ in range(5):
        cs.cell(row,2,None).fill=yfill
        cs.cell(row,3,None).fill=yfill
        for cl,col in [('B',4),('Q',5),('P',6),('T',7)]:
            cs.cell(row,col,lookup(cl,f'B{row}')).font=F(size=9)
        cs.cell(row,5).alignment=ctr; cs.cell(row,6).alignment=wrap
        cs.cell(row,8,None).fill=yfill; cs.cell(row,9,None).fill=yfill
        for c in range(1,10): cs.cell(row,c).border=border
        row+=1
    # widths + assign dropdown
    for col,w in zip('ABCDEFGHI',[6,24,52,9,9,30,12,11,26]):
        cs.column_dimensions[col].width=w
    dv2=DataValidation(type='list',formula1=assign_dv_formula,allow_blank=True)
    cs.add_data_validation(dv2); dv2.add(f'H{hrow+1}:H{row-1}')
    cs.freeze_panes='A6'

wb.save('/root/work/Journeys3_Mentor_Matching.xlsx')
print('saved. mentors:',len(mrows),' company tabs:',len(matches),' data rows end:',DATA_END,' adhoc start:',ADHOC_START)
